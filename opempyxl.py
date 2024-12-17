import openpyxl

#path="C:\SQL2019\cn.xlsx"
#workbook=openpyxl.load_workbook(path)
#sheet=workbook["Sheet1"]

#rows=sheet.max_row  #count numbers of rows in a excel sheet
#cols=sheet.max_column  #count numbers of col in a excel sheet


#Reading all the rows and columns in excel sheet

#for row in range(1,rows+1):
#    for col in range(1,cols+1):
#        print(sheet.cell(row,col).value)

#writing in excel
#path="C:\SQL2019\Test.xlsx"
#workbook=openpyxl.load_workbook(path)
#sheet=workbook.active --- for all the active sheets

#for r in range(1,6):
#    for c in range(1,4):
#        sheet.cell(r,c).value="welcome"

#workbook.save(path)

path="C:\SQL2019\Test.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet2"]

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

workbook.save(path)







